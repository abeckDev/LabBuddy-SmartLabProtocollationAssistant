"""
LabBuddy — Session storage: saves conversation artifacts locally
and optionally uploads to Azure Blob Storage.
"""

import io
import json
import logging
import os
import wave
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

logger = logging.getLogger(__name__)

EXPORTS_DIR = Path(__file__).parent / "exports"

# Default brand color for Excel header
_HEADER_FILL = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)


def _make_header_fill(brand_color: str = "#0078D4") -> PatternFill:
    """Create a PatternFill from a hex color string."""
    color = brand_color.lstrip("#")
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def pcm16_to_wav(pcm_bytes: bytes, sample_rate: int = 24000) -> bytes:
    """Wrap raw PCM16 mono bytes into a WAV container."""
    buf = io.BytesIO()
    with wave.open(buf, "wb") as wf:
        wf.setnchannels(1)
        wf.setsampwidth(2)
        wf.setframerate(sample_rate)
        wf.writeframes(pcm_bytes)
    return buf.getvalue()


class SessionStorage:
    """Manages export artifacts for a single conversation session."""

    def __init__(self, session_id: str, config: dict | None = None):
        self.session_id = session_id
        self.local_dir = EXPORTS_DIR / session_id
        self.local_dir.mkdir(parents=True, exist_ok=True)
        self.config = config
        self._header_fill = (
            _make_header_fill(config["brand_color"])
            if config and "brand_color" in config
            else _HEADER_FILL
        )

    async def save_artifacts(
        self,
        transcript: list[dict],
        audio_bytes: bytes,
        fields: dict,
    ) -> dict[str, str]:
        """Save all artifacts locally and optionally to blob storage.
        Returns dict mapping artifact name → download path.
        """
        self._save_transcript(transcript)
        self._save_wav(audio_bytes)
        self._save_json(fields)
        self._save_xlsx(fields)
        await self._upload_to_blob()

        base = f"/api/download/{self.session_id}"
        return {
            "transcript": f"{base}/transcript.txt",
            "audio": f"{base}/conversation.wav",
            "json": f"{base}/extracted_fields.json",
            "xlsx": f"{base}/extracted_fields.xlsx",
        }

    # -- Local file writers -----------------------------------------------

    def _save_transcript(self, transcript: list[dict]):
        path = self.local_dir / "transcript.txt"
        with open(path, "w", encoding="utf-8") as f:
            f.write(f"LabBuddy Experiment Protocol Transcript\n")
            f.write(f"Session: {self.session_id}\n")
            f.write(f"Date: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}\n")
            f.write("=" * 60 + "\n\n")
            for entry in transcript:
                ts = entry.get("timestamp", "")
                role = entry.get("role", "unknown").upper()
                text = entry.get("text", "")
                f.write(f"[{ts}] {role}:\n{text}\n\n")

    def _save_wav(self, pcm_bytes: bytes):
        if not pcm_bytes:
            return
        wav_data = pcm16_to_wav(pcm_bytes)
        path = self.local_dir / "conversation.wav"
        with open(path, "wb") as f:
            f.write(wav_data)

    def _save_json(self, fields: dict):
        path = self.local_dir / "extracted_fields.json"
        output = {
            "session_id": self.session_id,
            "exported_at": datetime.now(timezone.utc).isoformat(),
            "fields": fields,
        }
        with open(path, "w", encoding="utf-8") as f:
            json.dump(output, f, indent=2, ensure_ascii=False)

    def _save_xlsx(self, fields: dict):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "LabBuddy Protocol Extract"

        # Header row
        for col, header in enumerate(["Field", "Value"], 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = _HEADER_FONT
            cell.fill = self._header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for i, (key, value) in enumerate(fields.items(), start=2):
            ws.cell(row=i, column=1, value=key)
            ws.cell(row=i, column=2, value=str(value) if value is not None else "")

        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 55

        path = self.local_dir / "extracted_fields.xlsx"
        wb.save(path)

    # -- Azure Blob Storage (optional) ------------------------------------

    async def _upload_to_blob(self):
        endpoint = os.getenv("AZURE_STORAGE_ENDPOINT")
        if not endpoint:
            logger.info("No AZURE_STORAGE_ENDPOINT set — skipping blob upload")
            return

        try:
            from azure.identity.aio import DefaultAzureCredential
            from azure.storage.blob.aio import BlobServiceClient

            credential = DefaultAzureCredential()
            async with BlobServiceClient(endpoint, credential=credential) as client:
                container = client.get_container_client("conversations")
                # Ensure container exists
                try:
                    await container.create_container()
                except Exception:
                    pass  # already exists

                for file_path in self.local_dir.iterdir():
                    blob_name = f"{self.session_id}/{file_path.name}"
                    blob = container.get_blob_client(blob_name)
                    with open(file_path, "rb") as f:
                        await blob.upload_blob(f, overwrite=True)

                logger.info(f"Uploaded {self.session_id} artifacts to blob storage")
            await credential.close()
        except Exception as e:
            logger.warning(f"Blob upload failed (local files still available): {e}")
